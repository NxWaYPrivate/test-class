from flask import Flask, render_template, request, jsonify
import cv2
import numpy as np
import base64
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)

sessions = {}
excel_file = "sessions_travail.xlsx"

try:
    workbook = load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Nom QR", "Date début", "Pause (oui/non)", "Durée de pause (minutes)", "Heure de fin", "Temps de travail effectif"])
    workbook.save(excel_file)

@app.route('/')
def index():
    return render_template('html_test.html')

@app.route('/scan_base64', methods=['POST'])
def scan_qr_base64():
    data = request.get_json()
    image_data = data.get('image')
    if image_data:
        import sys
        print(f"Taille image base64 : {len(image_data)}", file=sys.stderr)
        try:
            encoded_data = image_data.split(',')[1]
            img_bytes = base64.b64decode(encoded_data)
            npimg = np.frombuffer(img_bytes, np.uint8)
            img = cv2.imdecode(npimg, cv2.IMREAD_COLOR)
            if img is None:
                raise ValueError("Image non décodable")
        except Exception as e:
            return jsonify({"error": f"Erreur décodage image : {str(e)}"}), 400

        detector = cv2.QRCodeDetector()
        data, _, _ = detector.detectAndDecode(img)

        if data:
            qr_name = data.strip()
            if qr_name not in sessions:
                sessions[qr_name] = {
                    "start_time": datetime.now(),
                    "pause": False,
                    "pause_duration": 0,
                    "end_time": None,
                }
            return jsonify({"message": f"QR Code détecté : {qr_name}"})
        else:
            return jsonify({"error": "Aucun QR Code détecté"}), 400
    return jsonify({"error": "Aucune image reçue"}), 400

@app.route('/pause', methods=['POST'])
def pause_session():
    data = request.get_json()
    qr_name = data.get("qr_name")
    if qr_name in sessions:
        session = sessions[qr_name]
        if not session["pause"]:
            session["pause"] = True
            session["pause_start"] = datetime.now()
            return jsonify({"message": "Pause commencée"})
        else:
            pause_start = session.pop("pause_start", datetime.now())
            pause_duration = (datetime.now() - pause_start).total_seconds() / 60
            session["pause_duration"] += pause_duration
            session["pause"] = False
            return jsonify({"message": f"Pause terminée. Durée ajoutée : {pause_duration:.2f} minutes"})
    return jsonify({"error": "Session non trouvée"}), 400

@app.route('/stop', methods=['POST'])
def stop_session():
    data = request.get_json()
    qr_name = data.get("qr_name")
    if qr_name in sessions:
        session = sessions[qr_name]
        session["end_time"] = datetime.now()

        total_work_time = (session["end_time"] - session["start_time"]).total_seconds() / 60
        effective_work_time = total_work_time - session["pause_duration"]

        workbook = load_workbook(excel_file)
        sheet = workbook.active
        sheet.append([
            qr_name,
            session["start_time"].strftime("%Y-%m-%d %H:%M:%S"),
            "Oui" if session["pause_duration"] > 0 else "Non",
            f"{round(session['pause_duration'], 2)} minutes",
            session["end_time"].strftime("%Y-%m-%d %H:%M:%S"),
            f"{round(effective_work_time, 2)} minutes"
        ])
        workbook.save(excel_file)

        sessions.pop(qr_name)
        return jsonify({"message": "Session terminée et enregistrée"}), 200
    return jsonify({"error": "Session non trouvée"}), 400

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5050)